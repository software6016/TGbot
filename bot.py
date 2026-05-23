import os
import io
import re
import time
import logging
import requests
import pandas as pd
from openpyxl import load_workbook
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    ApplicationBuilder,
    CommandHandler,
    MessageHandler,
    ContextTypes,
    filters,
)
from dotenv import load_dotenv

# Load environment variables from .env file
load_dotenv()

# ── Настройки ──────────────────────────────────────────────────────────────────

BOT_TOKEN          = os.getenv("TELEGRAM_BOT_TOKEN")
ONEDRIVE_SHARE_URL = os.getenv("ONEDRIVE_SHARE_URL")

if not BOT_TOKEN:
    raise ValueError("TELEGRAM_BOT_TOKEN environment variable is required")
if not ONEDRIVE_SHARE_URL:
    raise ValueError("ONEDRIVE_SHARE_URL environment variable is required")

CACHE_TTL = 60  # секунд — как часто обновлять файл из OneDrive
_cache: dict = {"result": None, "ts": 0.0}

# Внутренние имена колонок после нормализации
COL_TRACKING  = "_tracking"
COL_CLIENT    = "_client"
COL_DESC      = "_desc"
COL_SENT      = "_sent"
COL_METHOD    = "_method"
COL_WEIGHT    = "_weight"
COL_PRICE     = "_price"
COL_NOTES     = "_notes"
COL_CITY_CODE = "_city_code"  # Код города/направления (AE/SE/HE) — колонка I

# Тарифы ($/кг) по коду города и типу доставки
TARIFF_CARGO = {"AE": 3.2, "SE": 3.4, "HE": 3.6}
TARIFF_AVIA  = {"AE": 10.0, "SE": 11.5, "HE": 12.0}

# Контакты поддержки и каналы
WHATSAPP_URL = "https://wa.me/77474365528"
TG_MAIN_URL  = "https://t.me/flashgo6016kz"
TG_LOST_URL  = "https://t.me/bezkoda6016"

# ── Логирование ────────────────────────────────────────────────────────────────

logging.basicConfig(
    format="%(asctime)s │ %(levelname)s │ %(message)s",
    level=logging.INFO,
)
logger = logging.getLogger(__name__)

# ── Кнопки ─────────────────────────────────────────────────────────────────────

def support_keyboard() -> InlineKeyboardMarkup:
    """Inline-клавиатура с кнопками поддержки и каналов."""
    keyboard = [
        [InlineKeyboardButton("💬 Служба поддержки (WhatsApp)", url=WHATSAPP_URL)],
        [
            InlineKeyboardButton("📢 Канал 6016.kz", url=TG_MAIN_URL),
            InlineKeyboardButton("📦 Потеряшки", url=TG_LOST_URL),
        ],
    ]
    return InlineKeyboardMarkup(keyboard)


def tariff_keyboard() -> InlineKeyboardMarkup:
    """Кнопка под сообщением с тарифами."""
    keyboard = [
        [InlineKeyboardButton("💬 Задать вопрос (WhatsApp)", url=WHATSAPP_URL)],
    ]
    return InlineKeyboardMarkup(keyboard)

# ── Скачивание Excel из OneDrive ───────────────────────────────────────────────

def download_excel_bytes() -> bytes:
    """Скачивает Excel файл из OneDrive sharing-ссылки."""
    headers = {"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36"}

    r = requests.get(ONEDRIVE_SHARE_URL, headers=headers, timeout=30, allow_redirects=True)
    html = r.text
    logger.info("Страница просмотра: статус %s, размер %d", r.status_code, len(html))

    patterns = [
        r'"downloadUrl"\s*:\s*"([^"]+)"',
        r'"url"\s*:\s*"(https://[^"]*\.xlsx[^"]*)"',
        r'downloadUrl["\s:]+["\']?(https://[^\s"\'<>]+)',
        r'"FileGetUrl"\s*:\s*"([^"]+)"',
        r'sj\.u\(["\']([^"\']*download[^"\']*)["\']',
    ]

    download_url = None
    for pattern in patterns:
        match = re.search(pattern, html)
        if match:
            download_url = match.group(1).replace("\\u0026", "&").replace("\\/", "/")
            logger.info("Найдена download ссылка через паттерн: %s", pattern)
            break

    if download_url:
        r2 = requests.get(download_url, headers=headers, timeout=30, allow_redirects=True)
        ct = r2.headers.get("Content-Type", "")
        logger.info("Скачивание: статус %s, Content-Type: %s, размер: %d", r2.status_code, ct, len(r2.content))
        if r2.status_code == 200 and len(r2.content) > 5000:
            return r2.content

    logger.error("Download URL не найден. Кусок HTML:\n%s", html[:3000])
    raise RuntimeError("Не удалось найти прямую ссылку на скачивание в OneDrive")

# ── Зелёные строки (не прошли контроль) ───────────────────────────────────────

def is_row_green(ws, row_idx: int) -> bool:
    """True если строка закрашена зелёным (задержана / не прошла контроль)."""
    green = 0
    total = 0
    for cell in ws[row_idx]:
        if cell.value is None:
            continue
        total += 1
        fill = cell.fill
        if not fill or fill.fill_type != "solid":
            continue
        try:
            color = fill.fgColor
            if color.type == "rgb":
                rgb = str(color.rgb)
                if len(rgb) == 8 and rgb not in ("FFFFFFFF", "00000000", "FF000000", "00FFFFFF"):
                    r_val = int(rgb[2:4], 16)
                    g_val = int(rgb[4:6], 16)
                    b_val = int(rgb[6:8], 16)
                    if g_val > r_val and g_val > b_val and g_val > 60:
                        green += 1
            elif color.type == "theme":
                if color.theme in (6, 9):
                    green += 1
            elif color.type == "indexed":
                if color.indexed in (4, 10, 17, 50, 57):
                    green += 1
        except Exception:
            pass
    return total > 0 and (green / total) > 0.3

# ── Нормализация листов ────────────────────────────────────────────────────────

def normalize_sheet(df: pd.DataFrame, sheet_name: str) -> pd.DataFrame | None:
    """Нормализует структуру листа AVIA или CARGO по ПОЗИЦИОННЫМ индексам.

    Структура файла фиксирована и keyword-поиск заголовков ненадёжен
    (в новых AVIA-листах колонка трек-кода имеет пустой/нестандартный
    заголовок, из-за чего листы пропускались целиком). Поэтому
    используем жёсткие позиции:

    CARGO листы:
      A=0 №,  B=1 ДАТА,  C=2 ПУНКТ,    D=3 ТРЕК-КОД,
      E=4 город (AE/SE/HE…),  F=5 НАЗВАНИЕ,  G=6 ВЕС,
      H=7 $/кг,             I=8 ОПЛ В КЗ (итог),  J=9 КОММЕНТАРИЙ

    AVIA листы:
      A=0 №,  B=1 ДАТА,  C=2 внеш. упак.,  D=3 ПУНКТ,
      E=4 ТРЕК-КОД,       F=5 город (AE/SE/HE…),
      G=6 НАЗВАНИЕ,       H=7 МЕСТА,        I=8 ВЕС,
      J=9 货代重量,        K=10 НУЖНО ПОЛУЧИТЬ (итог),
      L=11 ИМЯ/НОМЕР,     M=12 ОПЛ В КЗ,    N=13 ОПЛ В КИТАЕ,
      O=14 КОММЕНТАРИЙ
    """
    df.columns = [str(c).strip() for c in df.columns]

    upper_name = sheet_name.upper()
    is_avia  = "AVIA"  in upper_name
    is_cargo = "CARGO" in upper_name
    if not (is_avia or is_cargo):
        return None

    method = "авиа" if is_avia else "наземная"

    # Позиционная раскладка для каждого типа листа
    if is_avia:
        idx_sent, idx_track, idx_city = 1, 4, 5
        idx_desc, idx_weight, idx_price = 6, 8, 10
        idx_client, idx_notes = 11, 14
    else:  # CARGO
        idx_sent, idx_track, idx_city = 1, 3, 4
        idx_desc, idx_weight, idx_price = 5, 6, 8
        idx_client, idx_notes = None, 9  # в CARGO имени получателя нет

    ncols = len(df.columns)
    n_rows = len(df)
    empty_series = pd.Series([""] * n_rows, index=df.index)

    def col_or_blank(i):
        if i is None or i >= ncols:
            return empty_series.copy()
        return df.iloc[:, i].astype(str)

    df[COL_METHOD]    = method
    df[COL_SENT]      = col_or_blank(idx_sent)
    df[COL_TRACKING]  = col_or_blank(idx_track)
    # Город = первые 2 заглавные буквы значения (AE6622 → AE, HE878103 → HE, 5020 → "").
    # Пустые/None ячейки сначала чистим, чтобы случайно не получить "NO" из "None".
    _city = col_or_blank(idx_city).str.strip().str.upper()
    _city = _city.replace({"NONE": "", "NAN": "", "NAT": ""})
    df[COL_CITY_CODE] = _city.str.extract(r"^([A-Z]{2})", expand=False).fillna("")
    df[COL_DESC]      = col_or_blank(idx_desc)
    df[COL_WEIGHT]    = col_or_blank(idx_weight)
    df[COL_PRICE]     = col_or_blank(idx_price)
    df[COL_CLIENT]    = col_or_blank(idx_client)
    df[COL_NOTES]     = col_or_blank(idx_notes)

    # Нормализуем трек-коды и убираем мусорные строки
    def _norm_track(val) -> str:
        s = str(val).strip()
        if s.endswith(".0") and s[:-2].isdigit():
            s = s[:-2]
        return s.upper()

    df[COL_TRACKING] = df[COL_TRACKING].apply(_norm_track)
    df = df[~df[COL_TRACKING].isin(["", "NAN", "NONE", "NAT"])]

    return df if not df.empty else None

# ── Работа с Excel ─────────────────────────────────────────────────────────────

def load_orders():
    """Скачивает Excel из OneDrive, читает ВСЕ листы, находит зелёные строки.
    При ошибке скачивания — возвращает устаревший кэш вместо падения.
    """
    now = time.time()
    if _cache["result"] and now - _cache["ts"] < CACHE_TTL:
        return _cache["result"]

    try:
        excel_bytes = download_excel_bytes()
    except Exception as e:
        logger.error("Ошибка скачивания Excel: %s", e)
        if _cache["result"]:
            logger.warning("Возвращаем устаревший кэш")
            return _cache["result"]
        return (pd.DataFrame(), set())

    buf = io.BytesIO(excel_bytes)

    xl = pd.ExcelFile(buf)
    logger.info("Листы в файле: %s", xl.sheet_names)
    frames = []

    for sheet in xl.sheet_names:
        try:
            df_raw = pd.read_excel(xl, sheet_name=sheet, header=1, dtype=str)
            df_norm = normalize_sheet(df_raw, sheet)
            if df_norm is not None and not df_norm.empty:
                df_norm["_sheet"] = sheet
                frames.append(df_norm)
                logger.info("✅ Лист «%s»: %d строк", sheet, len(df_norm))
            else:
                logger.warning("Лист «%s» — трек-колонка не найдена или пуста", sheet)
        except Exception as e:
            logger.warning("Лист «%s» пропущен: %s", sheet, e)

    if not frames:
        logger.error("Ни один лист не содержит трек-коды")
        result = (pd.DataFrame(), set())
        _cache["result"] = result
        _cache["ts"] = now
        return result

    df_all = pd.concat(frames, ignore_index=True)

    # Зелёные строки через openpyxl — позиционно, согласовано с normalize_sheet:
    #   AVIA → трек в колонке E (openpyxl 1-indexed = 5)
    #   CARGO → трек в колонке D (openpyxl 1-indexed = 4)
    # Заголовки — Excel row 2, данные начинаются с row 3.
    buf.seek(0)
    wb = load_workbook(buf, data_only=True)
    green_tracks: set[str] = set()

    def _norm_track_str(val) -> str:
        s = str(val).strip()
        if s.endswith(".0") and s[:-2].isdigit():
            s = s[:-2]
        return s.upper()

    for sheet_name in wb.sheetnames:
        upper_name = sheet_name.upper()
        is_avia  = "AVIA"  in upper_name
        is_cargo = "CARGO" in upper_name
        if not (is_avia or is_cargo):
            continue

        ws = wb[sheet_name]
        track_col_idx = 5 if is_avia else 4  # E or D (1-indexed)

        for row_idx in range(3, ws.max_row + 1):
            track_val = ws.cell(row=row_idx, column=track_col_idx).value
            if track_val and is_row_green(ws, row_idx):
                green_tracks.add(_norm_track_str(track_val))

    logger.info("Зелёных строк (не прошли контроль): %d", len(green_tracks))
    result = (df_all, green_tracks)
    _cache["result"] = result
    _cache["ts"] = now
    return result


def find_order(code: str):
    """Ищет заказ по трек-коду."""
    df, green_tracks = load_orders()
    code = code.strip().upper()
    if code.endswith(".0") and code[:-2].isdigit():
        code = code[:-2]
    matches = df[df[COL_TRACKING] == code]
    if matches.empty:
        matches = df[df[COL_TRACKING].str.contains(code, na=False, regex=False)]
    if matches.empty:
        return None, "not_found"

    is_green = code in green_tracks
    cargo_matches = matches[matches["_sheet"].str.upper().str.contains("CARGO")]
    avia_matches  = matches[matches["_sheet"].str.upper().str.contains("AVIA")]

    if is_green:
        if not cargo_matches.empty:
            return cargo_matches.iloc[0].to_dict(), "transferred_to_cargo"
        else:
            row = avia_matches.iloc[0] if not avia_matches.empty else matches.iloc[0]
            return row.to_dict(), "detained"
    else:
        result_row = cargo_matches.iloc[0] if not cargo_matches.empty else matches.iloc[0]
        return result_row.to_dict(), "ok"

# ── Форматирование ─────────────────────────────────────────────────────────────

def _year_from_sheet(sheet: str) -> int | None:
    """Достаёт год из имени листа: «MAY 2026 AVIA» → 2026."""
    if not sheet:
        return None
    m = re.search(r"(20\d{2})", sheet)
    return int(m.group(1)) if m else None


def _parse_short_date(s: str, year: int | None = None) -> pd.Timestamp | None:
    """Парсит «короткий» формат M.D / M.DD из Excel в pd.Timestamp.

    Excel хранит числовые ячейки как float и теряет хвостовые нули,
    поэтому в таблице действует конвенция:
      "5.01" / "5.05" / "5.09" → литеральный день 1-9 (с ведущим нулём)
      "5.1"  → день 10  (исходно "5.10", хвостовой 0 потерян)
      "5.2"  → день 20  (исходно "5.20")
      "5.3"  → день 30  (исходно "5.30")
      "5.5"  → день 5   (50 > 31 → конвенция не работает, литерал)
      "5.15" → день 15  (две цифры → литерал)
      "5.31" → день 31

    Возвращает None если строка не похожа на дату.
    """
    if not s:
        return None
    s = str(s).strip()
    m = re.match(r"^(\d{1,2})\.(\d{1,2})$", s)
    if not m:
        return None
    month = int(m.group(1))
    day_str = m.group(2)
    if len(day_str) == 1:
        d_int = int(day_str)
        # Одна цифра БЕЗ ведущего нуля → исходно был хвостовой 0 → десятки.
        # Но если *10 > 31 — это не валидный день → считаем литералом.
        day = d_int * 10 if d_int * 10 <= 31 else d_int
    else:
        day = int(day_str)
    if not (1 <= month <= 12 and 1 <= day <= 31):
        return None
    yr = year if year is not None else pd.Timestamp.now().year
    try:
        return pd.Timestamp(year=yr, month=month, day=day)
    except Exception:
        return None


def fmt_date(value, sheet: str = "") -> str:
    """Форматирует дату в DD.MM.YYYY. Год берём из имени листа, если возможно."""
    if pd.isna(value) or str(value).strip() in ("", "nan", "NaT"):
        return "—"
    s = str(value).strip()
    ts = _parse_short_date(s, year=_year_from_sheet(sheet))
    if ts is not None:
        return ts.strftime("%d.%m.%Y")
    try:
        return pd.Timestamp(value).strftime("%d.%m.%Y")
    except Exception:
        return s


def calc_arrival(sent_value, method: str, sheet: str = "") -> str:
    """Рассчитывает примерный диапазон дат прибытия."""
    try:
        if pd.isna(sent_value) or str(sent_value).strip() in ("", "nan", "NaT"):
            return "—"
        s = str(sent_value).strip()
        sent_ts = _parse_short_date(s, year=_year_from_sheet(sheet))
        if sent_ts is None:
            sent_ts = pd.Timestamp(sent_value)
        m = str(method).lower()
        if "авиа" in m or "air" in m:
            d1 = (sent_ts + pd.Timedelta(days=3)).strftime("%d.%m.%Y")
            d2 = (sent_ts + pd.Timedelta(days=4)).strftime("%d.%m.%Y")
        else:
            d1 = (sent_ts + pd.Timedelta(days=7)).strftime("%d.%m.%Y")
            d2 = (sent_ts + pd.Timedelta(days=12)).strftime("%d.%m.%Y")
        return f"{d1} — {d2}"
    except Exception:
        return "—"


def fmt_method(value) -> str:
    s = str(value).strip().lower()
    if "авиа" in s or "air" in s:
        return "✈️ авиа"
    if "наземн" in s or "ground" in s or "land" in s:
        return "🚚 наземная"
    return f"📦 {value}"


def get_val(order: dict, col: str) -> str:
    v = str(order.get(col, "")).strip()
    return "—" if v in ("", "nan", "None", "NaT") else v


def calc_price(order: dict) -> str:
    """Рассчитывает стоимость доставки: тариф по коду города × вес."""
    method     = get_val(order, COL_METHOD).lower()
    city_code  = get_val(order, COL_CITY_CODE).upper().strip()
    weight_str = get_val(order, COL_WEIGHT)

    rates = TARIFF_AVIA if ("авиа" in method or "air" in method) else TARIFF_CARGO
    rate  = rates.get(city_code)

    if rate is not None and weight_str != "—":
        try:
            weight = float(weight_str)
            price  = weight * rate
            return f"💰 *Стоимость:* ${price:,.2f}  (${rate}/кг × {weight} кг)"
        except Exception:
            pass

    # Запасной вариант — показываем сырое значение из Excel
    raw = get_val(order, COL_PRICE)
    return f"💰 *Стоимость:* {raw}"


def build_reply(order: dict, header: str | None = None) -> str:
    raw_method = get_val(order, COL_METHOD)
    method     = fmt_method(raw_method)
    sheet      = str(order.get("_sheet", ""))
    sent       = fmt_date(order.get(COL_SENT), sheet)
    arrival    = calc_arrival(order.get(COL_SENT), raw_method, sheet)
    client     = get_val(order, COL_CLIENT)
    desc       = get_val(order, COL_DESC)
    notes      = get_val(order, COL_NOTES)
    price_str  = calc_price(order)

    try:
        weight_str = f"⚖️ *Вес:* {float(get_val(order, COL_WEIGHT))} кг"
    except Exception:
        weight_str = f"⚖️ *Вес:* {get_val(order, COL_WEIGHT)}"

    title_line = header if header else f"📦 *Ваш товар отправлен* {sent}"

    lines = [
        title_line,
        f"Способ доставки: {method}",
        "",
        "────────────────────",
        f"🔖 *Трек-код:* `{get_val(order, COL_TRACKING)}`",
        f"👤 *Получатель:* {client}",
        f"📝 *Товар:* {desc}",
        price_str,
        weight_str,
        f"📅 *Дата отправки:* {sent}",
        f"🕐 *Примерная дата прибытия:* {arrival}",
    ]
    if notes != "—":
        lines.append(f"📌 *Примечание:* {notes}")

    lines += [
        "────────────────────",
        "_Если у вас есть вопросы — обратитесь в нашу службу поддержки!_ 👇",
    ]
    return "\n".join(lines)


def build_reply_transferred(order: dict) -> str:
    sheet = str(order.get("_sheet", ""))
    sent = fmt_date(order.get(COL_SENT), sheet)
    header = (
        "⚠️ *Ваша посылка не прошла авиа-контроль*\n\n"
        "Отправление было возвращено на склад и переведено на *наземную доставку (карго)*.\n\n"
        f"🚚 Новая дата отправки: *{sent}*"
    )
    return build_reply(order, header=header)

# ── Обработчики ────────────────────────────────────────────────────────────────

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    await update.message.reply_text(
        "👋 *Здравствуйте!*\n"
        "Вас приветствует Telegram-бот *6016.kz*.\n\n"
        "Чем могу быть полезен?",
        parse_mode="Markdown",
    )
    await update.message.reply_text(
        "📦 Пожалуйста, отправьте *трек-код* товара, чтобы получить информацию о доставке.\n\n"
        "Примеры трек-кодов:\n"
        "`SF1234567891011`\n"
        "`JDK1234567890`\n"
        "`1234567890`\n\n"
        "Поддерживаются все форматы курьерских служб.\n\n"
        "_Если у вас есть вопросы — мы всегда на связи!_ 👇",
        parse_mode="Markdown",
        reply_markup=support_keyboard(),
    )


async def help_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    await update.message.reply_text(
        "📦 *Как узнать статус доставки:*\n\n"
        "Просто отправьте ваш *трек-код* товара.\n\n"
        "Примеры:\n"
        "`SF1234567891011`\n"
        "`JDK1234567890`\n"
        "`1234567890`\n\n"
        "Трек-код можно найти в чеке или подтверждении заказа.\n\n"
        "🔄 Для принудительного обновления данных используйте /refresh\n"
        "📋 Тарифы на доставку — /tariff",
        parse_mode="Markdown",
        reply_markup=support_keyboard(),
    )


async def tariff_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Показывает тарифы на доставку."""
    await update.message.reply_text(
        "📋 *Тарифы на доставку:*\n\n"
        "🚚 *Наземная доставка (карго):*\n"
        "• AE — $3.2/кг\n"
        "• SE — $3.4/кг\n"
        "• HE — $3.6/кг\n\n"
        "✈️ *Авиадоставка:*\n"
        "• AE — $10/кг\n"
        "• SE — $11.5/кг\n"
        "• HE — $12/кг\n\n"
        "_По вопросам тарифов обращайтесь в службу поддержки_ 👇",
        parse_mode="Markdown",
        reply_markup=tariff_keyboard(),
    )


async def support_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Показывает кнопки службы поддержки."""
    await update.message.reply_text(
        "🤝 *Служба поддержки 6016.kz*\n\n"
        "Если у вас есть какие-либо вопросы — мы всегда рады помочь!\n\n"
        "Нажмите кнопку ниже для связи 👇",
        parse_mode="Markdown",
        reply_markup=support_keyboard(),
    )


async def refresh_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Принудительно сбрасывает кэш и перезагружает данные из Excel."""
    _cache["result"] = None
    _cache["ts"] = 0.0
    await update.message.reply_text(
        "🔄 Кэш сброшен. Данные будут обновлены при следующем запросе.",
        parse_mode="Markdown",
    )


async def track(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    user_input = update.message.text.strip()
    user = update.effective_user
    logger.info("Запрос от %s (%s): %r", user.id, user.first_name, user_input)

    await context.bot.send_chat_action(chat_id=update.effective_chat.id, action="typing")

    try:
        order, status = find_order(user_input)
    except Exception as e:
        logger.error("Ошибка при загрузке данных: %s", e)
        await update.message.reply_text(
            "⚠️ Не удалось загрузить данные. Попробуйте позже.",
            parse_mode="Markdown",
        )
        return

    if order is None:
        await update.message.reply_text(
            "⚠️ *Упс…*\n\n"
            "Похоже, товар ещё не был отправлен или не поступил на наш склад.\n\n"
            "🔍 Попробуйте проверить статус позже.\n\n"
            "_Если у вас есть вопросы — обратитесь в службу поддержки_ 👇",
            parse_mode="Markdown",
            reply_markup=support_keyboard(),
        )
    elif status == "transferred_to_cargo":
        await update.message.reply_text(
            "✅ *Товар найден!*\n\nСейчас проверяю информацию… ⏳",
            parse_mode="Markdown",
        )
        await context.bot.send_chat_action(chat_id=update.effective_chat.id, action="typing")
        await update.message.reply_text(
            build_reply_transferred(order),
            parse_mode="Markdown",
            reply_markup=support_keyboard(),
        )
    elif status == "detained":
        await update.message.reply_text(
            "✅ *Товар найден!*\n\nСейчас проверяю информацию… ⏳",
            parse_mode="Markdown",
        )
        await context.bot.send_chat_action(chat_id=update.effective_chat.id, action="typing")
        await update.message.reply_text(
            "⚠️ *Ваша посылка не прошла авиа-контроль.*\n\n"
            "Отправление находится на складе и готовится к переводу на *наземную доставку (карго)*.\n\n"
            "_По вопросам обращайтесь в нашу службу поддержки_ 👇",
            parse_mode="Markdown",
            reply_markup=support_keyboard(),
        )
    else:
        await update.message.reply_text(
            "✅ *Товар найден!*\n\nСейчас проверяю информацию о доставке… ⏳",
            parse_mode="Markdown",
        )
        await context.bot.send_chat_action(chat_id=update.effective_chat.id, action="typing")
        await update.message.reply_text(
            build_reply(order),
            parse_mode="Markdown",
            reply_markup=support_keyboard(),
        )

# ── Запуск ─────────────────────────────────────────────────────────────────────

def main() -> None:
    logger.info("Бот запускается…")
    app = ApplicationBuilder().token(BOT_TOKEN).build()
    app.add_handler(CommandHandler("start",   start))
    app.add_handler(CommandHandler("help",    help_cmd))
    app.add_handler(CommandHandler("refresh", refresh_cmd))
    app.add_handler(CommandHandler("tariff",  tariff_cmd))
    app.add_handler(CommandHandler("support", support_cmd))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, track))
    app.run_polling(allowed_updates=Update.ALL_TYPES)


if __name__ == "__main__":
    main()
